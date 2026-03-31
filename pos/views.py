from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count, Q, F
from django.views.decorators.http import require_POST
import json, decimal
from datetime import date, timedelta

from .models import Store, UserProfile, Product, Sale, SaleItem, StockLog, Expense, AreaManagerStore, WholesaleApproval, Employee, PaySlip


def today_range():
    """Return (start, end) datetime range for TODAY in local timezone — fixes UTC vs IST mismatch."""
    import datetime
    from django.utils.timezone import make_aware, get_current_timezone
    tz    = get_current_timezone()
    today = date.today()
    start = make_aware(datetime.datetime.combine(today, datetime.time.min), tz)
    end   = make_aware(datetime.datetime.combine(today, datetime.time.max), tz)
    return start, end


def date_range(d):
    """Return (start, end) datetime range for a given date in local timezone."""
    import datetime
    from django.utils.timezone import make_aware, get_current_timezone
    tz    = get_current_timezone()
    start = make_aware(datetime.datetime.combine(d, datetime.time.min), tz)
    end   = make_aware(datetime.datetime.combine(d, datetime.time.max), tz)
    return start, end


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def get_profile(user):
    try:
        return user.profile
    except UserProfile.DoesNotExist:
        if user.is_superuser:
            return UserProfile.objects.create(user=user, role='SUPERADMIN')
        return None


def require_profile(view_fn):
    """Decorator: user must have a UserProfile."""
    from functools import wraps
    @wraps(view_fn)
    def inner(request, *args, **kwargs):
        p = get_profile(request.user)
        if not p:
            from django.contrib.auth import logout
            logout(request)
            messages.error(request, 'Your account has no role assigned. Contact admin.')
            return redirect('login')
        # Re-evaluate expiry on every request for temporary roles
        if p.has_expired:
            from django.contrib.auth import logout
            request.session.flush()
            logout(request)
            messages.error(request, 'Your temporary access role has expired.')
            return redirect('login')
        return view_fn(request, *args, **kwargs)
    return inner


def store_for_request(request):
    """Return the store the current user belongs to (None for superadmin)."""
    p = get_profile(request.user)
    if p and p.is_superadmin:
        return None          # superadmin has no single store
    return p.store if p else None


def assert_store_access(profile, store):
    """Return True if user can access this store."""
    if profile.is_superadmin:
        return True
    return profile.store_id == store.id


# ══════════════════════════════════════════════════════════════════════════════
#  AUTH
# ══════════════════════════════════════════════════════════════════════════════
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        from django.core.cache import cache
        from .audit import log_event
        import time
        
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        ip       = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', ''))
        lock_key = f"login_attempts:{ip}:{username}"
        attempts = cache.get(lock_key, 0)

        if attempts >= 5:
            messages.error(request, 'Too many failed attempts. Try again in 30 minutes.')
            log_event(request, 'LOGIN_LOCKED', f'username={username}', level='WARNING')
            return render(request, 'pos/login.html')

        user = authenticate(request, username=username, password=password)
        print(f"[DEBUG] Login attempt: username={username}, password={password}, authenticated={user}")
        if user:
            print(f"[DEBUG] Profile exists: {bool(get_profile(user))}")
            if get_profile(user):
                print(f"[DEBUG] has_expired: {get_profile(user).has_expired}")
        if user and get_profile(user) and not get_profile(user).has_expired:
            cache.delete(lock_key)
            login(request, user)
            log_event(request, 'LOGIN_SUCCESS', f'username={user.username}')
            return redirect('dashboard')
        else:
            cache.set(lock_key, attempts + 1, timeout=1800)  # 30-min window
            time.sleep(0.3) # Throttle to slow down brute force
            log_event(request, 'LOGIN_FAILED', f'username={username}', level='WARNING')
            messages.error(request, 'Invalid username or password.')
            
    return render(request, 'pos/login.html')


def logout_view(request):
    from .audit import log_event
    if request.user.is_authenticated:
        log_event(request, 'LOGOUT', f'username={request.user.username}')
    request.session.flush()
    logout(request)
    return redirect('login')


# ══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
@login_required
@require_profile
def dashboard(request):
    profile = get_profile(request.user)
    today   = date.today()

    t_start, t_end = today_range()
    
    # Run the email reminders job once a day when the dashboard loads
    from django.core.cache import cache
    if not cache.get(f'credit_reminders_sent_{today.isoformat()}'):
        try:
            from django.core.management import call_command
            call_command('check_credits')
            cache.set(f'credit_reminders_sent_{today.isoformat()}', True, timeout=86400)
        except Exception:
            pass

    from .models import CreditRecord
    from datetime import timedelta
    target_date = today + timedelta(days=2)
    
    if profile.is_superadmin or profile.role in ('AREAMANAGER', 'WHOLESALE_EXEC'):
        # Global view across all stores
        if profile.is_superadmin:
            stores = Store.objects.filter(is_active=True)
        else:
            stores = Store.objects.filter(id__in=AreaManagerStore.objects.filter(manager__user=request.user).values('store_id'))
            
        store_data = []
        for s in stores:
            today_sales = Sale.objects.filter(store=s, created_at__range=(t_start, t_end))
            agg = SaleItem.objects.filter(
                sale__store=s, sale__created_at__range=(t_start, t_end)
            ).aggregate(sales=Sum('total_amount'), cost=Sum('total_cost'), profit=Sum('profit'))
            low_stock = Product.objects.filter(
                store=s, is_active=True,
                stock_quantity__lte=F('low_stock_alert'), stock_quantity__gt=0
            ).count()
            out_stock = Product.objects.filter(store=s, is_active=True, stock_quantity__lte=0).count()
            store_data.append({
                'store':        s,
                'bill_count':   today_sales.count(),
                'total_sales':  agg['sales']  or 0,
                'total_profit': agg['profit'] or 0,
                'low_stock':    low_stock,
                'out_stock':    out_stock,
            })
        global_agg = SaleItem.objects.filter(
            sale__created_at__range=(t_start, t_end)
        ).aggregate(sales=Sum('total_amount'), profit=Sum('profit'))
        urgent_credits = CreditRecord.objects.filter(is_paid=False, due_date__lte=target_date)
        if not profile.is_superadmin:
            urgent_credits = urgent_credits.filter(sale__store__in=stores)
        urgent_credits = urgent_credits.order_by('due_date')
        
        from django.core.cache import cache
        recent_otps = cache.get(f'am_dashboard_otps_{profile.id}', []) if profile.role in ('AREAMANAGER', 'WHOLESALE_EXEC') else []
        
        ctx = {
            'profile':       profile,
            'store_data':    store_data,
            'global_sales':  global_agg['sales']  or 0,
            'global_profit': global_agg['profit'] or 0,
            'total_stores':  stores.count(),
            'total_bills':   Sale.objects.filter(created_at__range=(t_start, t_end)).count(),
            'urgent_credits': urgent_credits,
            'recent_otps':   recent_otps,
        }
        return render(request, 'pos/dashboard_admin.html', ctx)

    else:
        # Store-scoped dashboard
        store = profile.store
        if not store:
            messages.error(request, 'You are not assigned to any store.')
            return render(request, 'pos/no_store.html', {'profile': profile})

        today_items = SaleItem.objects.filter(
            sale__store=store, sale__created_at__range=(t_start, t_end))
        agg = today_items.aggregate(
            sales=Sum('total_amount'), cost=Sum('total_cost'), profit=Sum('profit'))

        week_start_dt, _ = date_range(today - timedelta(days=6))
        week_items = SaleItem.objects.filter(
            sale__store=store, sale__created_at__gte=week_start_dt)
        week_agg = week_items.aggregate(sales=Sum('total_amount'), profit=Sum('profit'))

        low_products = Product.objects.filter(
            store=store, is_active=True,
            stock_quantity__lte=F('low_stock_alert'), stock_quantity__gt=0)
        out_products = Product.objects.filter(store=store, is_active=True, stock_quantity__lte=0)
        recent_sales = Sale.objects.filter(store=store).order_by('-created_at')[:8]

        urgent_credits = CreditRecord.objects.filter(
            Q(sale__store=store) | Q(is_external=True),
            is_paid=False, due_date__lte=target_date
        ).order_by('due_date')

        ctx = {
            'profile':      profile,
            'store':        store,
            'today_sales':  agg['sales']       or 0,
            'today_cost':   agg['cost']        or 0,
            'today_profit': agg['profit']      or 0,
            'today_bills':  Sale.objects.filter(store=store, created_at__range=(t_start, t_end)).count(),
            'week_sales':   week_agg['sales']  or 0,
            'week_profit':  week_agg['profit'] or 0,
            'low_products': low_products[:5],
            'low_count':    low_products.count(),
            'out_count':    out_products.count(),
            'recent_sales': recent_sales,
            'urgent_credits': urgent_credits,
        }
        return render(request, 'pos/dashboard_store.html', ctx)


# ══════════════════════════════════════════════════════════════════════════════
#  STORE MANAGEMENT  (superadmin only)
# ══════════════════════════════════════════════════════════════════════════════
@login_required
@require_profile
def store_list(request):
    profile = get_profile(request.user)
    if not profile.is_superadmin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    stores_qs = Store.objects.filter(is_active=True).order_by('name')
    # Annotate separately to avoid JOIN multiplication
    stores_with_staff    = {s.id: s.staff_count    for s in stores_qs.annotate(staff_count=Count('staff',    filter=Q(staff__is_active=True),    distinct=True))}
    stores_with_products = {s.id: s.product_count  for s in stores_qs.annotate(product_count=Count('products', filter=Q(products__is_active=True), distinct=True))}
    stores = list(stores_qs)
    for s in stores:
        s.staff_count   = stores_with_staff.get(s.id, 0)
        s.product_count = stores_with_products.get(s.id, 0)
    return render(request, 'pos/stores.html', {'stores': stores, 'profile': profile})


@login_required
@require_profile
def store_create(request):
    profile = get_profile(request.user)
    if not profile.is_superadmin:
        return redirect('dashboard')
    if request.method == 'POST':
        Store.objects.create(
            name=request.POST['name'].strip(),
            address=request.POST.get('address', '').strip(),
            phone=request.POST.get('phone', '').strip(),
            whatsapp_number=request.POST.get('whatsapp_number', '').strip(),
            email=request.POST.get('email', '').strip(),
            gstin=request.POST.get('gstin', '').strip(),
        )
        messages.success(request, f'Store "{request.POST["name"]}" created!')
    return redirect('store_list')


@login_required
@require_profile
def store_edit(request, store_id):
    profile = get_profile(request.user)
    if not profile.is_superadmin:
        return redirect('dashboard')
    store = get_object_or_404(Store, id=store_id)
    if request.method == 'POST':
        store.name            = request.POST.get('name', store.name).strip()
        store.address         = request.POST.get('address', store.address).strip()
        store.phone           = request.POST.get('phone', store.phone).strip()
        store.whatsapp_number = request.POST.get('whatsapp_number', store.whatsapp_number).strip()
        store.email           = request.POST.get('email', store.email).strip()
        store.gstin           = request.POST.get('gstin', store.gstin).strip()
        store.is_active       = request.POST.get('is_active') == 'on'
        store.save()
        messages.success(request, 'Store updated.')
    return redirect('store_list')


@login_required
@require_profile
def store_detail(request, store_id):
    """Admin view into a specific store."""
    profile = get_profile(request.user)
    if not profile.is_superadmin:
        return redirect('dashboard')
    store   = get_object_or_404(Store, id=store_id)
    today   = date.today()
    report_date_str = request.GET.get('date', '')
    try:
        report_date = date.fromisoformat(report_date_str) if report_date_str else today
    except ValueError:
        report_date = today

    r_start, r_end = date_range(report_date)

    items_qs = SaleItem.objects.filter(
        sale__store=store, sale__created_at__range=(r_start, r_end))
    agg = items_qs.aggregate(
        sales=Sum('total_amount'), cost=Sum('total_cost'), profit=Sum('profit'))

    # Month range
    import datetime
    from django.utils.timezone import make_aware, get_current_timezone
    tz = get_current_timezone()
    month_start = make_aware(datetime.datetime(report_date.year, report_date.month, 1, 0, 0, 0), tz)
    import calendar
    last_day = calendar.monthrange(report_date.year, report_date.month)[1]
    month_end = make_aware(datetime.datetime(report_date.year, report_date.month, last_day, 23, 59, 59), tz)
    month_agg = SaleItem.objects.filter(
        sale__store=store, sale__created_at__range=(month_start, month_end)
    ).aggregate(sales=Sum('total_amount'), profit=Sum('profit'))

    ctx = {
        'profile':       profile,
        'store':         store,
        'report_date':   report_date,
        'sales':         Sale.objects.filter(store=store, created_at__range=(r_start, r_end)),
        'sale_items':    items_qs,
        'total_sales':   agg['sales']         or 0,
        'total_cost':    agg['cost']          or 0,
        'total_profit':  agg['profit']        or 0,
        'month_sales':   month_agg['sales']   or 0,
        'month_profit':  month_agg['profit']  or 0,
        'products':      Product.objects.filter(store=store, is_active=True),
        'staff':         UserProfile.objects.filter(store=store, is_active=True).select_related('user'),
        'expenses':      Expense.objects.filter(store=store, date=report_date),
    }
    return render(request, 'pos/store_detail.html', ctx)


# ══════════════════════════════════════════════════════════════════════════════
#  USER MANAGEMENT  (superadmin only)
# ══════════════════════════════════════════════════════════════════════════════
@login_required
@require_profile
def user_management(request):
    profile = get_profile(request.user)
    if not profile.is_superadmin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    users  = UserProfile.objects.select_related('user', 'store').order_by('store__name', 'role')
    stores = Store.objects.filter(is_active=True)
    return render(request, 'pos/users.html', {'users': users, 'stores': stores, 'profile': profile})


@login_required
@require_profile
def user_create(request):
    profile = get_profile(request.user)
    if not profile.is_superadmin:
        return redirect('dashboard')
    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        password   = request.POST.get('password', '')
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        role       = request.POST.get('role', 'STAFF')
        store_id   = request.POST.get('store_id')

        if not username:
            messages.error(request, 'Username is required.')
        elif not email:
            messages.error(request, 'Email address is required.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, f'Username "{username}" is already taken.')
        elif User.objects.filter(email=email).exists():
            messages.error(request, f'Email "{email}" is already taken by another account.')
        else:
            ALLOWED_ROLES = ['SUBADMIN', 'OWNER', 'STAFF', 'AREAMANAGER', 'WHOLESALE_EXEC']
            if role not in ALLOWED_ROLES:
                messages.error(request, 'Invalid role assignment.')
                return redirect('user_management')
            if role == 'SUPERADMIN':
                messages.error(request, 'Cannot create Super Admin via this interface.')
                return redirect('user_management')
                
            from django.contrib.auth.password_validation import validate_password
            from django.core.exceptions import ValidationError
            from .audit import log_event
            
            try:
                validate_password(password)
            except ValidationError as e:
                for error in e.messages:
                    messages.error(request, error)
                return redirect('user_management')

            u = User.objects.create_user(username=username, password=password, email=email,
                                         first_name=first_name, last_name=last_name)
            log_event(request, 'USER_CREATED', f'created_user={username} role={role}')
            
            store = None
            if store_id and store_id != 'None':
                store = Store.objects.filter(id=store_id).first()
            phone = request.POST.get('phone', '').strip()

            expires_at_str = request.POST.get('expires_at', '').strip()
            expires_at = None
            if expires_at_str:
                import datetime
                from django.utils import timezone
                try:
                    expires_at = timezone.make_aware(datetime.datetime.strptime(expires_at_str, '%Y-%m-%dT%H:%M'))
                except Exception:
                    pass

            permissions = {}
            if role == 'SUBADMIN':
                permissions = {
                    'manage_users': request.POST.get('perm_manage_users') == 'on',
                    'view_reports': request.POST.get('perm_view_reports') == 'on',
                    'manage_inventory': request.POST.get('perm_manage_inventory') == 'on',
                }

            UserProfile.objects.create(
                user=u, role=role, store=store, phone=phone,
                expires_at=expires_at, permissions=permissions
            )
            role_label = dict(UserProfile.ROLE_CHOICES).get(role, role)
            messages.success(request, f'User "{username}" created as {role_label}.')
            if role in ('AREAMANAGER', 'WHOLESALE_EXEC'):
                messages.info(request, f'Go to Area Managers page to assign stores and set a PIN for {username}.')

    return redirect('user_management')


@login_required
@require_profile
def user_edit(request, user_id):
    profile = get_profile(request.user)
    if not profile.is_superadmin:
        return redirect('dashboard')
    up = get_object_or_404(UserProfile, id=user_id)
    if request.method == 'POST':
        up.role      = request.POST.get('role', up.role)
        store_id     = request.POST.get('store_id', '').strip()
        up.store     = Store.objects.filter(id=store_id).first() if (store_id and store_id != 'None') else None
        up.is_active = request.POST.get('is_active') == 'on'
        up.phone     = request.POST.get('phone', '').strip()

        expires_at_str = request.POST.get('expires_at', '').strip()
        if expires_at_str:
            import datetime
            from django.utils import timezone
            try:
                up.expires_at = timezone.make_aware(datetime.datetime.strptime(expires_at_str, '%Y-%m-%dT%H:%M'))
            except Exception:
                pass
        else:
            up.expires_at = None

        if up.role == 'SUBADMIN':
            up.permissions = {
                'manage_users': request.POST.get('perm_manage_users') == 'on',
                'view_reports': request.POST.get('perm_view_reports') == 'on',
                'manage_inventory': request.POST.get('perm_manage_inventory') == 'on',
            }

        new_email = request.POST.get('email', '').strip()
        if new_email and new_email != up.user.email:
            if User.objects.filter(email=new_email).exclude(id=up.user.id).exists():
                messages.error(request, f'Email "{new_email}" is already used by another account.')
                return redirect('user_management')
            up.user.email = new_email

        up.save()
        up.user.first_name = request.POST.get('first_name', up.user.first_name).strip()
        up.user.last_name  = request.POST.get('last_name',  up.user.last_name).strip()
        up.user.is_active  = up.is_active
        new_password = request.POST.get('password', '').strip()
        if new_password:
            from django.contrib.auth.password_validation import validate_password
            from django.core.exceptions import ValidationError
            try:
                validate_password(new_password)
                up.user.set_password(new_password)
            except ValidationError as e:
                for error in e.messages:
                    messages.error(request, error)
                return redirect('user_management')

        up.user.save()
        messages.success(request, f'User "{up.user.username}" updated successfully.')

    return redirect('user_management')


@require_POST
@login_required
@require_profile
def user_delete(request, user_id):
    profile = get_profile(request.user)
    if not profile.is_superadmin:
        return redirect('dashboard')
    up = get_object_or_404(UserProfile, id=user_id)
    if up.user == request.user:
        messages.error(request, "Can't delete your own account.")
    else:
        from .audit import log_event
        del_user = up.user.username
        up.user.delete()
        log_event(request, 'USER_DELETED', f'deleted_user={del_user}')
        messages.success(request, 'User deleted.')
    return redirect('user_management')


# ══════════════════════════════════════════════════════════════════════════════
#  BILLING
# ══════════════════════════════════════════════════════════════════════════════
@login_required
@require_profile
def billing(request):
    profile = get_profile(request.user)
    store   = profile.store
    if not store:
        messages.error(request, 'Not assigned to a store.')
        return redirect('dashboard')
    products = Product.objects.filter(store=store, is_active=True).order_by('category', 'name')
    from .models import WholesaleCustomer
    w_customers = WholesaleCustomer.objects.all()
    return render(request, 'pos/billing.html', {
        'products': products, 
        'store': store, 
        'profile': profile,
        'wholesale_customers': w_customers
    })


from django.views.decorators.http import require_POST
from django_ratelimit.decorators import ratelimit

@require_POST
@login_required
@require_profile
@ratelimit(key='ip', rate='10/m', block=True)
def save_bill(request):
    profile = get_profile(request.user)
    store   = profile.store
    if not store:
        return JsonResponse({'success': False, 'error': 'Not assigned to a store.'})

    # JSON Bombing Prevention (1MB max body)
    if len(request.body) > 1024 * 1024:
        return JsonResponse({'success': False, 'error': 'Payload too large.'}, status=400)

    try:
        data       = json.loads(request.body)
        items_data = data.get('items', [])
        bill_type  = data.get('bill_type', 'RETAIL')
        payment    = data.get('payment_mode', 'CASH')
        
        # Helper for strictly validating numeric bounds
        def validated_decimal(val, min_val=0, max_val=9999999):
            try:
                d = decimal.Decimal(str(val))
                if d < min_val or d > max_val:
                    raise ValueError
                return d
            except:
                raise ValueError("Invalid numeric value provided.")
                
        gst_rate = validated_decimal(data.get('gst_rate', 0), 0, 100)
        discount = validated_decimal(data.get('discount', 0), 0, 9999999)

        if not items_data:
            return JsonResponse({'success': False, 'error': 'No items in bill.'})

        # Validate products belong to this store & check stock
        validated = []
        for it in items_data:
            product = get_object_or_404(Product, id=it['product_id'], store=store)
            qty = validated_decimal(it['quantity'], 0.001, 99999)
            
            # CRITICAL FIX: Do not trust client's selling price, fetch from DB
            if bill_type == 'WHOLESALE':
                sp = product.wholesale_price
            else:
                sp = product.retail_price
                
            if product.stock_quantity < qty:
                return JsonResponse({'success': False,
                    'error': f'Insufficient stock for {product.name}. Available: {product.stock_quantity} kg'})
            validated.append((product, qty, sp))

        # Build Sale
        sale = Sale(store=store, bill_type=bill_type, payment_mode=payment,
                    gst_rate=gst_rate, discount=discount, created_by=request.user)
        
        cname = data.get('customer_name', '').strip()
        cphone = data.get('customer_phone', '').strip()
        
        wc = None
        if bill_type == 'WHOLESALE':
            sale.customer_name  = cname
            sale.customer_phone = cphone
            sale.customer_gst   = data.get('customer_gst', '').strip()

            from .models import WholesaleCustomer, CreditRecord
            if cname:
                wc = WholesaleCustomer.objects.filter(name__iexact=cname).first()
                if payment == 'CREDIT':
                    if not wc:
                        wc = WholesaleCustomer.objects.create(
                            name=cname, phone=sale.customer_phone, gst=sale.customer_gst,
                            is_credit_enabled=True, credit_duration_days=7, created_by=request.user
                        )
                    else:
                        if not wc.is_credit_enabled:
                            return JsonResponse({'success': False, 'error': f'Credit is disabled for {cname}.'})
                        if CreditRecord.objects.filter(customer=wc, is_paid=False).exists():
                            return JsonResponse({'success': False, 'error': f'{cname} has active unpaid credits. Settle them first.'})
                
                sale.wholesale_customer = wc
        else:
            # Capture retail customer info too (for WhatsApp billing)
            sale.customer_name  = cname
            sale.customer_phone = cphone


        subtotal = sum(q * sp for _, q, sp in validated)
        sale.subtotal = subtotal - discount

        if bill_type == 'WHOLESALE' and gst_rate > 0:
            half             = gst_rate / decimal.Decimal('2')
            sale.cgst_amount = (sale.subtotal * half / 100).quantize(decimal.Decimal('0.01'))
            sale.sgst_amount = sale.cgst_amount
            sale.total_gst   = sale.cgst_amount + sale.sgst_amount
            sale.grand_total = sale.subtotal + sale.total_gst
        else:
            sale.grand_total = sale.subtotal

        sale.save()
        
        # Audit Logging
        from .audit import log_event
        log_event(request, 'BILL_CREATED', f'bill={sale.bill_number} amount={sale.grand_total} mode={payment}')

        if payment == 'CREDIT' and wc:
            from datetime import timedelta
            CreditRecord.objects.create(
                customer=wc, sale=sale,
                due_date=date.today() + timedelta(days=wc.credit_duration_days)
            )

        # Create SaleItems + deduct stock
        for product, qty, sp in validated:
            cp = product.cost_price
            SaleItem.objects.create(
                sale=sale, product=product, product_name=product.name,
                quantity=qty, cost_price=cp, selling_price=sp,
                total_amount=qty*sp, total_cost=qty*cp, profit=qty*(sp-cp)
            )
            product.stock_quantity -= qty
            product.save(update_fields=['stock_quantity'])
            StockLog.objects.create(
                store=store, product=product, movement='OUT',
                quantity=qty, balance=product.stock_quantity,
                reference=sale.bill_number, created_by=request.user
            )

        # Build WhatsApp message & URL
        import urllib.parse
        items_summary = "\n".join(
            f"  • {p.name}: {q}kg × ₹{sp} = ₹{q*sp:.2f}"
            for p, q, sp in validated
        )
        wa_msg = (
            f"🌊 *Ocean Waves Sea Foods*\n"
            f"📍 {store.name}\n"
            f"──────────────────────\n"
            f"🧾 Bill No: *{sale.bill_number}*\n"
            f"📅 Date: {sale.created_at.strftime('%d/%m/%Y %I:%M %p')}\n"
            f"💳 Payment: {sale.get_payment_mode_display()}\n"
            f"──────────────────────\n"
            f"{items_summary}\n"
            f"──────────────────────\n"
        )
        if sale.discount > 0:
            wa_msg += f"🏷️ Discount: -₹{sale.discount:.2f}\n"
        if bill_type == 'WHOLESALE' and sale.total_gst > 0:
            wa_msg += f"🏛️ GST ({sale.gst_rate}%): ₹{sale.total_gst:.2f}\n"
        wa_msg += f"💰 *TOTAL: ₹{sale.grand_total:.2f}*\n"
        wa_msg += f"──────────────────────\n"
        wa_msg += f"✨ Thank you for shopping with us!\nFresh Seafood Every Day 🐟"
        
        # Build whatsapp URL — send to customer if they gave their number,
        # otherwise fall back to the store's own WhatsApp number
        customer_phone = sale.customer_phone.strip().lstrip('+').replace(' ', '').replace('-', '')
        store_wa = (store.whatsapp_number or '').strip().lstrip('+').replace(' ', '').replace('-', '')
        wa_phone = customer_phone if customer_phone else store_wa
        encoded_msg = urllib.parse.quote(wa_msg)
        whatsapp_url = f"https://wa.me/{wa_phone}?text={encoded_msg}" if wa_phone else None
        
        return JsonResponse({
            'success': True,
            'bill_id': sale.id,
            'bill_number': sale.bill_number,
            'whatsapp_url': whatsapp_url,
        })

    except Exception as e:
        from .audit import log_event
        log_event(request, 'BILL_API_ERROR', str(e), level='ERROR')
        # Mask exact internal errors
        return JsonResponse({'success': False, 'error': 'An internal error occurred while saving the bill.'})


@login_required
@require_profile
def bill_print(request, bill_id):
    profile = get_profile(request.user)
    sale    = get_object_or_404(Sale, id=bill_id)
    if not assert_store_access(profile, sale.store):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    return render(request, 'pos/bill_print.html', {'sale': sale})


# ══════════════════════════════════════════════════════════════════════════════
#  INVENTORY
# ══════════════════════════════════════════════════════════════════════════════
@login_required
@require_profile
def inventory(request):
    profile  = get_profile(request.user)
    store    = profile.store
    if not store:
        return redirect('dashboard')
    products = Product.objects.filter(store=store, is_active=True)
    return render(request, 'pos/inventory.html', {'products': products, 'store': store, 'profile': profile})


@login_required
@require_profile
def product_add(request):
    profile = get_profile(request.user)
    store   = profile.store
    if not store:
        return redirect('dashboard')
    if request.method == 'POST':
        p = Product(
            store=store,
            name=request.POST['name'].strip(),
            category=request.POST.get('category', 'FISH'),
            barcode=request.POST.get('barcode', '').strip(),
            cost_price=request.POST['cost_price'],
            retail_price=request.POST['retail_price'],
            wholesale_price=request.POST['wholesale_price'],
            stock_quantity=request.POST.get('stock_quantity', 0),
            low_stock_alert=request.POST.get('low_stock_alert', 5),
        )
        p.save()
        if float(p.stock_quantity) > 0:
            StockLog.objects.create(store=store, product=p, movement='IN',
                quantity=p.stock_quantity, balance=p.stock_quantity,
                reference='Initial stock', created_by=request.user)
        messages.success(request, f'Product "{p.name}" added.')
    return redirect('inventory')


@login_required
@require_profile
def product_edit(request, pid):
    profile = get_profile(request.user)
    p = get_object_or_404(Product, id=pid, store=profile.store)
    if request.method == 'POST':
        p.name            = request.POST.get('name', p.name).strip()
        p.category        = request.POST.get('category', p.category)
        p.cost_price      = request.POST.get('cost_price', p.cost_price)
        p.retail_price    = request.POST.get('retail_price', p.retail_price)
        p.wholesale_price = request.POST.get('wholesale_price', p.wholesale_price)
        p.low_stock_alert = request.POST.get('low_stock_alert', p.low_stock_alert)
        p.save()
        messages.success(request, f'"{p.name}" updated.')
    return redirect('inventory')


@login_required
@require_profile
def product_restock(request, pid):
    profile = get_profile(request.user)
    p = get_object_or_404(Product, id=pid, store=profile.store)
    if request.method == 'POST':
        qty = decimal.Decimal(request.POST.get('add_quantity', 0))
        p.stock_quantity += qty
        p.save(update_fields=['stock_quantity'])
        StockLog.objects.create(store=profile.store, product=p, movement='IN',
            quantity=qty, balance=p.stock_quantity,
            reference=request.POST.get('note', 'Restock'),
            created_by=request.user)
        messages.success(request, f'Added {qty} kg to {p.name}. New stock: {p.stock_quantity} kg')
    return redirect('inventory')


@login_required
@require_profile
def product_delete(request, pid):
    profile = get_profile(request.user)
    p = get_object_or_404(Product, id=pid, store=profile.store)
    p.is_active = False
    p.save()
    messages.success(request, f'"{p.name}" removed.')
    return redirect('inventory')


@login_required
@require_profile
def stock_log(request):
    profile = get_profile(request.user)
    store   = profile.store
    if not store:
        return redirect('dashboard')
    logs = StockLog.objects.filter(store=store).select_related('product', 'created_by')[:200]
    return render(request, 'pos/stock_log.html', {'logs': logs, 'store': store, 'profile': profile})


# ══════════════════════════════════════════════════════════════════════════════
#  REPORTS  (owner + superadmin)
# ══════════════════════════════════════════════════════════════════════════════
@login_required
@require_profile
def reports(request):
    profile = get_profile(request.user)
    if profile.is_staff_role:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    store   = profile.store
    report_date_str = request.GET.get('date', '')
    try:
        report_date = date.fromisoformat(report_date_str) if report_date_str else date.today()
    except ValueError:
        report_date = date.today()

    r_start, r_end = date_range(report_date)

    qs_filter = {'sale__created_at__range': (r_start, r_end)}
    if store:
        qs_filter['sale__store'] = store

    items = SaleItem.objects.filter(**qs_filter).select_related('sale', 'sale__store', 'product')
    agg   = items.aggregate(
        total_sales=Sum('total_amount'), total_cost=Sum('total_cost'), total_profit=Sum('profit'))

    sale_filter = {'created_at__range': (r_start, r_end)}
    if store:
        sale_filter['store'] = store
    sales = Sale.objects.filter(**sale_filter).prefetch_related('items').select_related('store')

    pay_breakdown  = sales.values('payment_mode').annotate(
        count=Count('id'), total=Sum('grand_total')).order_by('-total')
    type_breakdown = sales.values('bill_type').annotate(
        count=Count('id'), total=Sum('grand_total')).order_by('-total')

    expenses_filter = {'date': report_date}
    if store:
        expenses_filter['store'] = store
    expenses     = Expense.objects.filter(**expenses_filter)
    total_expense = expenses.aggregate(t=Sum('amount'))['t'] or 0

    ctx = {
        'profile':         profile,
        'store':           store,
        'report_date':     report_date,
        'sales':           sales,
        'sale_items':      items,
        'total_sales':     agg['total_sales']  or 0,
        'total_cost':      agg['total_cost']   or 0,
        'total_profit':    agg['total_profit'] or 0,
        'pay_breakdown':   pay_breakdown,
        'type_breakdown':  type_breakdown,
        'expenses':        expenses,
        'total_expense':   total_expense,
        'net_profit':      (agg['total_profit'] or 0) - total_expense,
    }
    return render(request, 'pos/reports.html', ctx)


@login_required
@require_profile
def export_excel(request):
    profile = get_profile(request.user)
    if profile.is_staff_role:
        return HttpResponse('Access denied', status=403)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    store = profile.store
    report_date_str = request.GET.get('date', '')
    try:
        report_date = date.fromisoformat(report_date_str) if report_date_str else date.today()
    except ValueError:
        report_date = date.today()

    r_start, r_end = date_range(report_date)
    qs_filter = {'sale__created_at__range': (r_start, r_end)}
    if store:
        qs_filter['sale__store'] = store
    items = SaleItem.objects.filter(**qs_filter).select_related('sale', 'sale__store')

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = f"Sales {report_date}"

    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', fgColor='0077B6')
    ctr       = Alignment(horizontal='center', vertical='center')
    thin      = Side(style='thin', color='CCCCCC')
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:J1')
    ws['A1'] = 'Ocean Waves Sea Foods — Sales Report'
    ws['A1'].font = Font(bold=True, size=14, color='0077B6')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:J2')
    ws['A2'] = f'Date: {report_date.strftime("%d %B %Y")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(italic=True)

    headers = ['Store', 'Bill No', 'Bill Type', 'Product', 'Qty (kg)', 'Mode',
               'CP/kg (₹)', 'SP/kg (₹)', 'Total (₹)', 'Profit (₹)']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, ctr, bdr

    row = 5
    for item in items:
        vals = [
            item.sale.store.name, item.sale.bill_number,
            item.sale.get_bill_type_display(), item.product_name,
            float(item.quantity), item.sale.get_payment_mode_display(),
            float(item.cost_price), float(item.selling_price),
            float(item.total_amount), float(item.profit),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = bdr
            cell.alignment = Alignment(horizontal='right' if c > 4 else 'left')
        row += 1

    totals = items.aggregate(s=Sum('total_amount'), p=Sum('profit'))
    tf = PatternFill('solid', fgColor='E9F5FF')
    ws.cell(row=row, column=1, value='TOTALS').font = Font(bold=True)
    ws.cell(row=row, column=1).fill = tf
    ws.cell(row=row, column=9, value=float(totals['s'] or 0)).font = Font(bold=True)
    ws.cell(row=row, column=9).fill = tf
    ws.cell(row=row, column=10, value=float(totals['p'] or 0)).font = Font(bold=True)
    ws.cell(row=row, column=10).fill = tf

    for i, w in enumerate([18,16,12,22,10,10,12,12,14,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="OceanWaves_Sales_{report_date}.xlsx"'
    wb.save(resp)
    return resp



# ══════════════════════════════════════════════════════════════════════════════
#  EXPENSES  (add / delete / PDF upload)
# ══════════════════════════════════════════════════════════════════════════════
@require_POST
@login_required
@require_profile
def expense_add(request):
    profile = get_profile(request.user)
    if not profile.is_owner:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    store = profile.store
    
    exp = Expense(
        store=store,
        category=request.POST.get('category', 'OTHER'),
        description=request.POST.get('description', '').strip(),
        amount=request.POST.get('amount', 0),
        date=request.POST.get('date', date.today()),
        created_by=request.user,
    )
    
    bill = request.FILES.get('bill_pdf')
    if bill:
        import magic
        file_magic = magic.from_buffer(bill.read(2048), mime=True)
        bill.seek(0)
        
        allowed_mimes = ['application/pdf', 'image/jpeg', 'image/png']
        if file_magic not in allowed_mimes:
            messages.error(request, 'Invalid file type. Only PDF, JPG, PNG are allowed.')
            return redirect('expenses_page')
            
        exp.bill_pdf = bill
        
    exp.save()
    messages.success(request, 'Expense recorded.')
    return redirect('expenses_page')


@require_POST
@login_required
@require_profile
def expense_delete(request, expense_id):
    profile = get_profile(request.user)
    if not profile.is_owner:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    exp = get_object_or_404(Expense, id=expense_id, store=profile.store)
    import os
    if exp.bill_pdf and hasattr(exp.bill_pdf, 'path'):
        try:
            if os.path.exists(exp.bill_pdf.path):
                os.remove(exp.bill_pdf.path)
        except Exception:
            pass
            
    from .audit import log_event
    log_event(request, 'EXPENSE_DELETED', f'expense={exp.id} amount={exp.amount}')
    exp.delete()
    messages.success(request, 'Expense deleted.')
    return redirect('expenses_page')


@login_required
@require_profile
def expenses_page(request):
    profile = get_profile(request.user)
    if not profile.is_owner:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    store = profile.store
    report_date_str = request.GET.get('date', '')
    try:
        report_date = date.fromisoformat(report_date_str) if report_date_str else date.today()
    except ValueError:
        report_date = date.today()
    expenses     = Expense.objects.filter(store=store)
    total        = expenses.aggregate(t=Sum('amount'))['t'] or 0
    day_expenses = expenses.filter(date=report_date)
    day_total    = day_expenses.aggregate(t=Sum('amount'))['t'] or 0
    return render(request, 'pos/expenses.html', {
        'profile':     profile, 'store': store,
        'expenses':    expenses.order_by('-date', '-created_at')[:100],
        'day_expenses': day_expenses,
        'report_date': report_date,
        'total':       total,
        'day_total':   day_total,
    })


# ══════════════════════════════════════════════════════════════════════════════
#  EMPLOYEE MANAGER
# ══════════════════════════════════════════════════════════════════════════════
@login_required
@require_profile
def employee_list(request):
    profile = get_profile(request.user)
    if not profile.is_owner:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    store     = profile.store
    employees = Employee.objects.filter(store=store).order_by('is_active', 'full_name')
    total_payroll = Employee.objects.filter(store=store, is_active=True).aggregate(
        t=Sum('basic_salary'))['t'] or 0
    return render(request, 'pos/employees.html', {
        'profile': profile, 'store': store,
        'employees': employees, 'total_payroll': total_payroll,
    })


@login_required
@require_profile
def employee_add(request):
    profile = get_profile(request.user)
    if not profile.is_owner:
        return redirect('dashboard')
    store = profile.store
    if request.method == 'POST':
        last   = Employee.objects.filter(store=store).order_by('-id').first()
        num    = (last.id + 1) if last else 1
        emp_id = f"EMP{store.id}{str(num).zfill(4)}"
        Employee.objects.create(
            store=store, employee_id=emp_id,
            full_name=request.POST.get('full_name', '').strip(),
            phone=request.POST.get('phone', '').strip(),
            email=request.POST.get('email', '').strip(),
            designation=request.POST.get('designation', '').strip(),
            employment_type=request.POST.get('employment_type', 'FULLTIME'),
            pay_cycle=request.POST.get('pay_cycle', 'MONTHLY'),
            basic_salary=request.POST.get('basic_salary', 0),
            allowances=request.POST.get('allowances', 0),
            deductions=request.POST.get('deductions', 0),
            date_joined=request.POST.get('date_joined', date.today()),
            notes=request.POST.get('notes', '').strip(),
            created_by=request.user,
        )
        messages.success(request, 'Employee added.')
    return redirect('employee_list')


@login_required
@require_profile
def employee_edit(request, emp_id):
    profile = get_profile(request.user)
    if not profile.is_owner:
        return redirect('dashboard')
    emp = get_object_or_404(Employee, id=emp_id, store=profile.store)
    if request.method == 'POST':
        emp.full_name       = request.POST.get('full_name', emp.full_name).strip()
        emp.phone           = request.POST.get('phone', emp.phone).strip()
        emp.email           = request.POST.get('email', emp.email).strip()
        emp.designation     = request.POST.get('designation', emp.designation).strip()
        emp.employment_type = request.POST.get('employment_type', emp.employment_type)
        emp.pay_cycle       = request.POST.get('pay_cycle', emp.pay_cycle)
        emp.basic_salary    = request.POST.get('basic_salary', emp.basic_salary)
        emp.allowances      = request.POST.get('allowances', emp.allowances)
        emp.deductions      = request.POST.get('deductions', emp.deductions)
        emp.is_active       = request.POST.get('is_active') == 'on'
        emp.notes           = request.POST.get('notes', emp.notes).strip()
        emp.save()
        messages.success(request, f'"{emp.full_name}" updated.')
    return redirect('employee_list')


@login_required
@require_profile
def employee_delete(request, emp_id):
    profile = get_profile(request.user)
    if not profile.is_owner:
        return redirect('dashboard')
    emp = get_object_or_404(Employee, id=emp_id, store=profile.store)
    emp.is_active = False
    emp.save()
    messages.success(request, f'"{emp.full_name}" deactivated.')
    return redirect('employee_list')


@login_required
@require_profile
def employee_detail(request, emp_id):
    profile  = get_profile(request.user)
    if not profile.is_owner:
        return redirect('dashboard')
    emp      = get_object_or_404(Employee, id=emp_id, store=profile.store)
    payslips = PaySlip.objects.filter(employee=emp).order_by('-year', '-month')
    return render(request, 'pos/employee_detail.html', {
        'profile': profile, 'emp': emp,
        'payslips': payslips, 'store': profile.store,
    })


@login_required
@require_profile
def payslip_generate(request, emp_id):
    profile = get_profile(request.user)
    if not profile.is_owner:
        return redirect('dashboard')
    emp = get_object_or_404(Employee, id=emp_id, store=profile.store)
    if request.method == 'POST':
        month = int(request.POST.get('month'))
        year  = int(request.POST.get('year'))
        if PaySlip.objects.filter(employee=emp, month=month, year=year).exists():
            messages.error(request, 'Payslip for this month already exists.')
        else:
            PaySlip.objects.create(
                employee=emp, store=profile.store,
                month=month, year=year,
                basic_salary=request.POST.get('basic_salary', emp.basic_salary),
                allowances=request.POST.get('allowances', emp.allowances),
                deductions=request.POST.get('deductions', emp.deductions),
                bonus=request.POST.get('bonus', 0),
                status=request.POST.get('status', 'PENDING'),
                payment_date=request.POST.get('payment_date') or None,
                payment_mode=request.POST.get('payment_mode', '').strip(),
                notes=request.POST.get('notes', '').strip(),
                created_by=request.user,
            )
            messages.success(request, f'Payslip generated for {emp.full_name}.')
    return redirect('employee_detail', emp_id=emp_id)


@login_required
@require_profile
def payslip_mark_paid(request, slip_id):
    profile = get_profile(request.user)
    if not profile.is_owner:
        return redirect('dashboard')
    slip = get_object_or_404(PaySlip, id=slip_id, store=profile.store)
    slip.status       = 'PAID'
    slip.payment_date = date.today()
    slip.payment_mode = request.POST.get('payment_mode', 'CASH')
    slip.save()
    messages.success(request, 'Payslip marked as paid.')
    return redirect('employee_detail', emp_id=slip.employee.id)


@login_required
@require_profile
def payslip_delete(request, slip_id):
    profile = get_profile(request.user)
    if not profile.is_owner:
        return redirect('dashboard')
    slip   = get_object_or_404(PaySlip, id=slip_id, store=profile.store)
    emp_id = slip.employee.id
    slip.delete()
    messages.success(request, 'Payslip deleted.')
    return redirect('employee_detail', emp_id=emp_id)


@login_required
@require_profile
def payslip_print(request, slip_id):
    profile = get_profile(request.user)
    slip    = get_object_or_404(PaySlip, id=slip_id, store=profile.store)
    return render(request, 'pos/payslip_print.html', {'slip': slip})


# ══════════════════════════════════════════════════════════════════════════════
#  API
# ══════════════════════════════════════════════════════════════════════════════
@login_required
@require_profile
def product_api(request, pid):
    profile = get_profile(request.user)
    p = get_object_or_404(Product, id=pid, store=profile.store, is_active=True)
    return JsonResponse({
        'id': p.id, 'name': p.name,
        'retail_price':    str(p.retail_price),
        'wholesale_price': str(p.wholesale_price),
        'stock_quantity':  str(p.stock_quantity),
        'category':        p.category,
    })


# ══════════════════════════════════════════════════════════════════════════════
#  AREA MANAGER MANAGEMENT  (superadmin only)
# ══════════════════════════════════════════════════════════════════════════════
@login_required
@require_profile
def area_manager_list(request):
    profile = get_profile(request.user)
    if not profile.is_superadmin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    managers  = UserProfile.objects.filter(role='AREAMANAGER').select_related('user').prefetch_related('managed_stores__store')
    stores    = Store.objects.filter(is_active=True)
    all_links = AreaManagerStore.objects.select_related('manager__user', 'store').order_by('store__name')
    return render(request, 'pos/area_managers.html', {
        'managers':  managers,
        'stores':    stores,
        'all_links': all_links,
        'profile':   profile,
    })


@login_required
@require_profile
def area_manager_assign(request):
    """Assign or remove a store from an area manager."""
    profile = get_profile(request.user)
    if not profile.is_superadmin:
        return redirect('dashboard')

    if request.method == 'POST':
        action     = request.POST.get('action')          # 'assign' or 'remove'
        manager_id = request.POST.get('manager_id')
        store_ids  = request.POST.getlist('store_ids')   # for assign: list of stores

        manager_profile = get_object_or_404(UserProfile, id=manager_id, role='AREAMANAGER')

        if action == 'assign':
            added = 0
            for sid in store_ids:
                store = get_object_or_404(Store, id=sid)
                _, created = AreaManagerStore.objects.get_or_create(
                    manager=manager_profile, store=store,
                    defaults={'assigned_by': request.user}
                )
                if created:
                    added += 1
            messages.success(request, f'Assigned {added} store(s) to {manager_profile.user.username}.')

        elif action == 'remove':
            link_id = request.POST.get('link_id')
            AreaManagerStore.objects.filter(id=link_id).delete()
            messages.success(request, 'Store assignment removed.')

        elif action == 'remove_all':
            AreaManagerStore.objects.filter(manager=manager_profile).delete()
            messages.success(request, f'All store assignments removed for {manager_profile.user.username}.')

    return redirect('area_manager_list')


# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
#  WHOLESALE PIN APPROVAL  (fast, zero-cost, works offline)
# ══════════════════════════════════════════════════════════════════════════════

def _hash_pin(raw_pin):
    from django.contrib.auth.hashers import make_password
    return make_password(raw_pin)

def _check_pin(profile, entered_pin):
    from django.contrib.auth.hashers import check_password
    if not profile.approval_pin:
        return False, 'PIN not set for this manager.'
    return check_password(str(entered_pin), profile.approval_pin), ''


@login_required
@require_profile
def wholesale_managers(request):
    """
    AJAX: Return list of Area Managers (with PINs set) for this store.
    Called when cashier opens the approval dialog — no bill data yet.
    """
    profile = get_profile(request.user)
    store   = profile.store
    if not store:
        return JsonResponse({'success': False, 'error': 'Not assigned to a store.'})

    am_links = AreaManagerStore.objects.filter(
        store=store, manager__is_active=True
    ).select_related('manager__user')

    if not am_links.exists():
        return JsonResponse({
            'success': False,
            'no_managers': True,
            'error': 'No Area Manager assigned to this store. Contact admin.'
        })

    managers_ready = []
    managers_no_email = []
    for link in am_links:
        am   = link.manager
        name = am.user.get_full_name() or am.user.username
        if am.user.email:
            managers_ready.append({'id': am.id, 'name': name})
        else:
            managers_no_email.append(name)

    return JsonResponse({
        'success':         True,
        'managers':        managers_ready,
        'managers_no_pin': managers_no_email,
    })


@login_required
@require_profile
def wholesale_request_otp(request):
    """
    AJAX: Generate 6-digit OTP, send it to the chosen manager, cache it.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'}, status=405)

    profile = get_profile(request.user)
    store   = profile.store
    if not store:
        return JsonResponse({'success': False, 'error': 'Not assigned to a store.'})

    try:
        data       = json.loads(request.body)
        manager_id = data.get('manager_id')

        if not manager_id:
            return JsonResponse({'success': False, 'error': 'Manager ID required.'})

        link = AreaManagerStore.objects.filter(
            store=store, manager__id=manager_id, manager__is_active=True
        ).select_related('manager__user').first()

        if not link:
            return JsonResponse({'success': False, 'error': 'Manager not found or not assigned to this store.'})

        am = link.manager
        email = am.user.email
        if not email:
            return JsonResponse({'success': False, 'error': 'Selected Manager has no email address configured.'})
            
        import random, string
        from django.core.mail import send_mail
        from django.conf import settings
        from django.core.cache import cache
        
        otp = ''.join(random.choices(string.digits, k=6))
        cache_key = f'ws_otp_{am.id}_{store.id}'
        cache.set(cache_key, otp, timeout=300) # 5 min

        # Surface OTP directly to AM Dashboard to bypass any SMTP failures
        am_dashboard_key = f'am_dashboard_otps_{am.id}'
        existing = cache.get(am_dashboard_key, [])
        import time, datetime
        now = time.time()
        existing = [o for o in existing if now - o['ts'] < 300]
        existing.insert(0, {
            'store_name': store.name,
            'time': datetime.datetime.now().strftime("%I:%M %p"),
            'code': otp,
            'ts': now
        })
        cache.set(am_dashboard_key, existing, timeout=300)

        subject = f"Wholesale Billing OTP - {store.name}"
        message = f"Hello {am.user.first_name},\n\nA wholesale bill is pending approval at {store.name}.\nYour OTP is: {otp}\n\nValid for 5 minutes."
        
        # Dispatch SMTP via thread as a backup
        import threading
        def send_otp_email(sub, msg, frm, to):
            try:
                send_mail(sub, msg, frm, [to], fail_silently=False)
            except Exception:
                pass
                
        threading.Thread(target=send_otp_email, args=(subject, message, settings.DEFAULT_FROM_EMAIL, email)).start()
        
        return JsonResponse({'success': True, 'msg': f'OTP dispatched & visible on Manager Dashboard'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_profile
def wholesale_verify_otp(request):
    """
    AJAX: Cashier submits bill data + chosen manager ID + entered OTP.
    Verifies OTP → saves bill → returns bill id for print.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'}, status=405)

    profile = get_profile(request.user)
    store   = profile.store
    if not store:
        return JsonResponse({'success': False, 'error': 'Not assigned to a store.'})

    try:
        data       = json.loads(request.body)
        manager_id = data.get('manager_id')
        entered_otp = str(data.get('pin', '')).strip()

        if not manager_id or not entered_otp:
            return JsonResponse({'success': False, 'error': 'Manager and OTP required.'})

        # Verify manager is assigned to this store
        link = AreaManagerStore.objects.filter(
            store=store, manager__id=manager_id, manager__is_active=True
        ).select_related('manager__user').first()

        if not link:
            return JsonResponse({'success': False, 'error': 'Manager not found.'})

        am = link.manager
        
        from django.core.cache import cache
        cache_key = f'ws_otp_{am.id}_{store.id}'
        cached_otp = cache.get(cache_key)
        
        if not cached_otp:
            return JsonResponse({
                'success': False,
                'wrong_pin': True,
                'error': 'OTP expired or not requested. Please request a new OTP.'
            })
            
        if cached_otp != entered_otp:
            return JsonResponse({
                'success': False,
                'wrong_pin': True,
                'error': f'Invalid OTP. Please check the email sent to the manager.'
            })

        # ── OTP correct — clear cache, save bill ──
        cache.delete(cache_key)

        items_data  = data.get('items', [])
        bill_type   = data.get('bill_type', 'WHOLESALE')
        payment     = data.get('payment_mode', 'CASH')
        gst_rate    = decimal.Decimal(str(data.get('gst_rate', 0)))
        discount    = decimal.Decimal(str(data.get('discount', 0)))

        # Re-validate stock
        validated = []
        for it in items_data:
            product = get_object_or_404(Product, id=it['product_id'], store=store)
            qty     = decimal.Decimal(str(it['quantity']))
            sp      = decimal.Decimal(str(it['selling_price']))
            if product.stock_quantity < qty:
                return JsonResponse({'success': False,
                    'error': f'Insufficient stock: {product.name} only has {product.stock_quantity} kg.'})
            validated.append((product, qty, sp))

        # Build Sale
        sale = Sale(
            store=store, bill_type=bill_type, payment_mode=payment,
            gst_rate=gst_rate, discount=discount, created_by=request.user
        )
        cname = data.get('customer_name',  '').strip()
        sale.customer_name  = cname
        sale.customer_phone = data.get('customer_phone', '').strip()
        sale.customer_gst   = data.get('customer_gst',   '').strip()

        wc = None
        from .models import WholesaleCustomer, CreditRecord
        if cname:
            wc = WholesaleCustomer.objects.filter(name__iexact=cname).first()
            if payment == 'CREDIT':
                if not wc:
                    wc = WholesaleCustomer.objects.create(
                        name=cname, phone=sale.customer_phone, gst=sale.customer_gst,
                        is_credit_enabled=True, credit_duration_days=7, created_by=request.user
                    )
                else:
                    if not wc.is_credit_enabled:
                        return JsonResponse({'success': False, 'error': f'Credit is disabled for {cname}.'})
                    if CreditRecord.objects.filter(customer=wc, is_paid=False).exists():
                        return JsonResponse({'success': False, 'error': f'{cname} has active unpaid credits. Settle them first.'})
            sale.wholesale_customer = wc

        subtotal      = sum(q * sp for _, q, sp in validated)
        sale.subtotal = subtotal - discount

        if gst_rate > 0:
            half             = gst_rate / 2
            sale.cgst_amount = (sale.subtotal * half / 100).quantize(decimal.Decimal('0.01'))
            sale.sgst_amount = sale.cgst_amount
            sale.total_gst   = sale.cgst_amount + sale.sgst_amount
            sale.grand_total = sale.subtotal + sale.total_gst
        else:
            sale.grand_total = sale.subtotal

        sale.save()
        
        if payment == 'CREDIT' and wc:
            from datetime import timedelta
            CreditRecord.objects.create(
                customer=wc, sale=sale,
                due_date=date.today() + timedelta(days=wc.credit_duration_days)
            )

        # Create SaleItems + deduct stock + log
        for product, qty, sp in validated:
            cp = product.cost_price
            SaleItem.objects.create(
                sale=sale, product=product, product_name=product.name,
                quantity=qty, cost_price=cp, selling_price=sp,
                total_amount=qty * sp, total_cost=qty * cp, profit=qty * (sp - cp)
            )
            product.stock_quantity -= qty
            product.save(update_fields=['stock_quantity'])
            StockLog.objects.create(
                store=store, product=product, movement='OUT',
                quantity=qty, balance=product.stock_quantity,
                reference=sale.bill_number, created_by=request.user
            )

        # Write approval audit record
        items_summary = ', '.join(
            f"{v[0].name} {v[1]}kg" for v in validated[:4]
        ) + (f' +{len(validated)-4} more' if len(validated) > 4 else '')

        WholesaleApproval.objects.create(
            store=store,
            area_manager=am,
            sale=sale,
            bill_snapshot={
                'items_summary': items_summary,
                'grand_total':   str(sale.grand_total),
                'bill_type':     bill_type,
                'customer':      sale.customer_name,
            },
            approved_by_name=am.user.get_full_name() or am.user.username,
            created_by=request.user,
        )

        # Build WhatsApp URL for wholesale bill
        import urllib.parse
        items_wa = "\n".join(
            f"  \u2022 {v[0].name}: {v[1]}kg \u00d7 \u20b9{v[2]} = \u20b9{v[1]*v[2]:.2f}"
            for v in validated
        )
        wa_msg = (
            f"\U0001f30a *Ocean Waves Sea Foods*\n"
            f"\U0001f4cd {store.name}\n"
            f"\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
            f"\U0001f9fe Bill No: *{sale.bill_number}*\n"
            f"\U0001f4c5 Date: {sale.created_at.strftime('%d/%m/%Y %I:%M %p')}\n"
            f"\U0001f3db\ufe0f *GST TAX INVOICE*\n"
            f"\U0001f4b3 Payment: {sale.get_payment_mode_display()}\n"
            f"\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
            f"{items_wa}\n"
            f"\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
        )
        if sale.discount > 0:
            wa_msg += f"\U0001f3f7\ufe0f Discount: -\u20b9{sale.discount:.2f}\n"
        if sale.total_gst > 0:
            wa_msg += f"GST ({sale.gst_rate}%): \u20b9{sale.total_gst:.2f}\n"
        wa_msg += f"\U0001f4b0 *TOTAL: \u20b9{sale.grand_total:.2f}*\n"
        wa_msg += f"\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
        wa_msg += f"\u2728 Thank you for your business!\nFresh Seafood Every Day \U0001f41f"

        customer_phone = sale.customer_phone.strip().lstrip('+').replace(' ', '').replace('-', '')
        store_wa = (store.whatsapp_number or '').strip().lstrip('+').replace(' ', '').replace('-', '')
        wa_phone = customer_phone if customer_phone else store_wa
        encoded_msg = urllib.parse.quote(wa_msg)
        whatsapp_url = f"https://wa.me/{wa_phone}?text={encoded_msg}" if wa_phone else None

        return JsonResponse({
            'success':       True,
            'bill_id':       sale.id,
            'bill_number':   sale.bill_number,
            'approved_by':   am.user.get_full_name() or am.user.username,
            'whatsapp_url':  whatsapp_url,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ── PIN Management (superadmin sets PINs for Area Managers) ──────────────────
@login_required
@require_profile
def set_manager_pin(request):
    """Superadmin sets or resets an Area Manager's approval PIN."""
    profile = get_profile(request.user)
    if not profile.is_superadmin:
        return JsonResponse({'success': False, 'error': 'Access denied.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'}, status=405)

    try:
        data       = json.loads(request.body)
        manager_id = data.get('manager_id')
        new_pin    = str(data.get('pin', '')).strip()

        if not new_pin.isdigit() or not (4 <= len(new_pin) <= 6):
            return JsonResponse({'success': False, 'error': 'PIN must be 4 to 6 digits.'})

        am = get_object_or_404(UserProfile, id=manager_id, role='AREAMANAGER')
        am.approval_pin = _hash_pin(new_pin)
        am.save(update_fields=['approval_pin'])

        return JsonResponse({
            'success': True,
            'message': f'PIN set for {am.user.get_full_name() or am.user.username}.'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ── Approval history view ────────────────────────────────────────────────────
@login_required
@require_profile
def approval_history(request):
    """Show wholesale approval log for owner/superadmin."""
    profile = get_profile(request.user)
    if profile.is_staff_role:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    store = profile.store
    qs    = WholesaleApproval.objects.select_related(
        'area_manager__user', 'store', 'sale', 'created_by'
    )
    if store:
        qs = qs.filter(store=store)
    qs = qs[:200]

    return render(request, 'pos/approval_history.html', {
        'approvals': qs,
        'store':     store,
        'profile':   profile,
    })

# ─────────────────────────────────────────────────────────────────────────────
#  WHOLESALE CUSTOMERS & CREDITS
# ─────────────────────────────────────────────────────────────────────────────
from .models import WholesaleCustomer, CreditRecord

@login_required
@require_profile
def wholesale_customers(request):
    profile = get_profile(request.user)
    customers = WholesaleCustomer.objects.all().order_by('-created_at')
    return render(request, 'pos/wholesale_customers.html', {'customers': customers, 'profile': profile})

@login_required
@require_profile
def wholesale_customer_add(request):
    profile = get_profile(request.user)
    if not (profile.is_superadmin or profile.is_wholesale_exec or profile.is_owner or profile.is_area_manager):
        messages.error(request, 'Access denied.')
        return redirect('wholesale_customers')
        
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if WholesaleCustomer.objects.filter(name__iexact=name).exists():
            messages.error(request, 'Customer with this name already exists.')
        else:
            WholesaleCustomer.objects.create(
                name=name,
                phone=request.POST.get('phone', ''),
                email=request.POST.get('email', ''),
                gst=request.POST.get('gst', ''),
                credit_duration_days=int(request.POST.get('credit_duration_days', 7)),
                is_credit_enabled=request.POST.get('is_credit_enabled') == 'on',
                created_by=request.user
            )
            messages.success(request, 'Customer added.')
    return redirect('wholesale_customers')

@login_required
@require_profile
def wholesale_customer_edit(request, cid):
    profile = get_profile(request.user)
    if not (profile.is_superadmin or profile.is_wholesale_exec or profile.is_owner or profile.is_area_manager):
        messages.error(request, 'Access denied.')
        return redirect('wholesale_customers')
        
    c = get_object_or_404(WholesaleCustomer, id=cid)
    if request.method == 'POST':
        c.name = request.POST.get('name', c.name)
        c.phone = request.POST.get('phone', c.phone)
        c.email = request.POST.get('email', c.email)
        c.gst = request.POST.get('gst', c.gst)
        c.credit_duration_days = int(request.POST.get('credit_duration_days', c.credit_duration_days))
        c.is_credit_enabled = request.POST.get('is_credit_enabled') == 'on'
        c.save()
        messages.success(request, 'Customer updated.')
    return redirect('wholesale_customers')

@login_required
@require_profile
def credits_list(request):
    profile = get_profile(request.user)
    records = CreditRecord.objects.select_related('customer', 'sale', 'sale__store').order_by('is_paid', 'due_date')
    
    if profile.store and not profile.is_wholesale_exec and not profile.is_area_manager and not profile.is_superadmin:
        records = records.filter(Q(sale__store=profile.store) | Q(is_external=True))

    return render(request, 'pos/credits.html', {'records': records, 'profile': profile})

@login_required
@require_profile
def credit_add_external(request):
    profile = get_profile(request.user)
    from .models import WholesaleCustomer, CreditRecord
    import datetime

    if request.method == 'POST':
        customer_id = request.POST.get('customer')
        amount      = request.POST.get('amount')
        due_date    = request.POST.get('due_date')
        ref         = request.POST.get('reference', '').strip()
        
        if not customer_id or not amount or not due_date:
            messages.error(request, "Please fill all required fields.")
            return redirect('credit_add_external')
            
        customer = get_object_or_404(WholesaleCustomer, id=customer_id)
        
        try:
            amt = decimal.Decimal(str(amount))
            CreditRecord.objects.create(
                customer=customer,
                amount=amt,
                is_external=True,
                external_reference=ref,
                due_date=datetime.datetime.strptime(due_date, "%Y-%m-%d").date(),
                is_paid=False
            )
            messages.success(request, f"External credit of ₹{amt} added for {customer.name}.")
            return redirect('credits_list')
        except Exception as e:
            messages.error(request, f"Error saving credit: {e}")
            return redirect('credit_add_external')
            
    customers = WholesaleCustomer.objects.filter(is_credit_enabled=True)
    return render(request, 'pos/credit_add_external.html', {'customers': customers, 'profile': profile})


@require_POST
@login_required
@require_profile
def credit_pay(request, cid):
    from .models import AreaManagerStore
    profile = get_profile(request.user)
    record = get_object_or_404(CreditRecord, id=cid)
    
    can_access = False
    if profile.is_superadmin:
        can_access = True
    elif profile.store and profile.store == record.sale.store:
        can_access = True
    elif profile.is_area_manager or profile.is_wholesale_exec:
        if AreaManagerStore.objects.filter(manager=profile, store=record.sale.store).exists():
            can_access = True
            
    if not can_access:
        messages.error(request, 'Access denied to this record.')
        from .audit import log_event
        log_event(request, 'IDOR_ATTEMPT', f'credit_pay record={cid}', level='WARNING')
        return redirect('credits_list')

    record.is_paid = True
    record.paid_on = timezone.now().date()
    record.save()
    messages.success(request, f'Credit for {record.customer.name} marked as paid.')
        
    return redirect('credits_list')
